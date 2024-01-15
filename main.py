from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from selenium.webdriver.support.ui import Select
from time import sleep
from threading import Thread
import json
from openpyxl import Workbook
import re

def Find_Element(driver : webdriver.Chrome, by, value : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(by, value)
            break
        except:
            pass
        sleep(0.1)
    return element

def Find_Elements(driver : webdriver.Chrome, by, value : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(by, value)
            if len(elements) > 0:
                break
        except:
            pass
        sleep(0.1)
    return elements

def Send_Keys(element : WebElement, content : str):
    element.clear()
    for i in content:
        element.send_keys(i)
        sleep(0.1)

wb = Workbook()
sheet = wb.active
item = ['link', 'Nome Fantasia', 'Telefone', 'Razao Social', 'CNPJ', 'Capital Social', 'Tipo', 'Porte', 'Atividade Principal', 'Situação', 'Data da Situação', 'Data Abertura', 'Endereço', 'Numero', 'Complemento', 'Cidade', 'Bairro', 'Estado', 'Cep', 'Natureza Juridica', 'Atividade Principal - Cnae', 'Atividade Secundaria - Cnae', 'Atividades secundarrias Textos', 'Quadro Societário', 'Email']

for i in range(1, 26):
    sheet.cell(row = 1, column = i).value = item[i-1]

service = Service(executable_path = "C:\chromedriver-win64\chromedriver.exe")
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9030")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.6099.217 Safari/537.3")
driver = webdriver.Chrome(options = options, service = service)

with open('total.json', 'r') as file:
    links = json.load(file)

start_row = 2
for item_index, item in enumerate(links):
    print(f'{item_index} --> {item["link"]}')
    driver.get(item["link"])
    sheet.cell(row = start_row, column = 1).value = item["link"]
    try:
        check_nomes = driver.find_elements(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[1]/div')
        for check_nome in check_nomes:
            find_nome_title = check_nome.text.split('\n')
            if find_nome_title[0] == "Nome Fantasia":
                nome_fantasia = find_nome_title[1]
                print(f'nome_fantasia : {nome_fantasia}')
                sheet.cell(row = start_row, column = 2).value = nome_fantasia
    except:
        pass
    try:
        telefone = driver.find_element(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[3]/div[1]/p[2]/a').text.replace(' ', '').replace('-', '')
        print(f'telefone : {telefone}')
        sheet.cell(row = start_row, column = 3).value = telefone
    except:
        pass
    try:
        check_razaos = driver.find_elements(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[1]/div')
        for check_razao in check_razaos:
            find_razao_title = check_razao.text.split('\n')
            if find_razao_title[0] == "Razão Social":
                razao = find_razao_title[1]
                print(f'razao : {razao}')
                sheet.cell(row = start_row, column = 4).value = razao
    except:
        pass
    try:
        check_cnpjs = driver.find_elements(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[1]/div')
        for check_cnpj in check_cnpjs:
            find_cnpj_title = check_cnpj.text.split('\n')
            if find_cnpj_title[0] == "CNPJ":
                cnpj = find_cnpj_title[1].replace('.', '').replace('/', '').replace('-', '')
                print(f'cnpj : {cnpj}')
                sheet.cell(row = start_row, column = 5).value = cnpj
    except:
        pass
    try:
        check_capitals = driver.find_elements(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[1]/div')
        for check_capital in check_capitals:
            find_capital_title = check_capital.text.split('\n')
            if find_capital_title[0] == "Capital Social":
                capital = find_capital_title[1].split(' ')
                print(f'capital : {capital[1]}')
                sheet.cell(row = start_row, column = 6).value = capital[1]
    except:
        pass
    try:
        check_tipos = driver.find_elements(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[1]/div')
        for check_tipo in check_tipos:
            find_tipo_title = check_tipo.text.split('\n')
            if find_tipo_title[0] == "Tipo":
                tipo = find_tipo_title[1]
                print(f'tipo : {tipo}')
                sheet.cell(row = start_row, column = 7).value = tipo
    except:
        pass
        # porte = driver.find_element(By.XPATH, '')
        # sheet.cell(row = start_row, column = 8).value = porte
    try:
        atividade_principal = driver.find_element(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[5]/div[1]/p[2]').text.split(' - ')
        print(f'atividade principal : {" - ".join(atividade_principal[1:])}')
        sheet.cell(row = start_row, column = 9).value = " - ".join(atividade_principal[1:])
    except:
        pass
    try:
        check_situacaos = driver.find_elements(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[1]/div')
        for check_situacao in check_situacaos:
            find_situacao_title = check_situacao.text.split('\n')
            if find_situacao_title[0] == "Situação Cadastral":
                situacao = find_situacao_title[1]
                print(f'situacao : {situacao}')
                sheet.cell(row = start_row, column = 10).value = situacao
    except:
        pass
    try:
        check_data_situacaos = driver.find_elements(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[1]/div')
        for check_data_situacao in check_data_situacaos:
            find_data_situacao_title = check_data_situacao.text.split('\n')
            if find_data_situacao_title[0] == "Data da Situação Cadastral":
                data_situacao = find_data_situacao_title[1]
                print(f'date_situacao : {data_situacao}')
                sheet.cell(row = start_row, column = 11).value = data_situacao
    except:
        pass
    try:
        check_data_aberturas = driver.find_elements(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[1]/div')
        for check_data_abertura in check_data_aberturas:
            find_data_abertura_title = check_data_abertura.text.split('\n')
            if find_data_abertura_title[0] == "Data Abertura":
                data_abertura = find_data_abertura_title[1]
                print(f'date_abertura : {data_abertura}')
                sheet.cell(row = start_row, column = 12).value = data_abertura
    except:
        pass
    try:
        endereco = driver.find_element(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[2]/div[1]/p[2]').text
        print(f'endereco : {endereco}')
        sheet.cell(row = start_row, column = 13).value = endereco
    except:
        pass
    try:
        numero = driver.find_element(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[2]/div[2]/p[2]').text
        print(f'numero : {numero}')
        sheet.cell(row = start_row, column = 14).value = numero
    except:
        pass
    try:
        complemento = driver.find_element(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[2]/div[3]/p[2]').text
        print(f'complemento : {complemento}')
        sheet.cell(row = start_row, column = 15).value = complemento
    except:
        pass
    cidade = 'ITAPIRA'
    sheet.cell(row = start_row, column = 16).value = cidade
    try:
        bairro = driver.find_element(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[2]/div[5]/p[2]').text
        print(f'bairro : {bairro}')
        sheet.cell(row = start_row, column = 17).value = bairro
    except:
        pass
    estado = 'SP'
    sheet.cell(row = start_row, column = 18).value = estado
    try:
        cep = driver.find_element(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[2]/div[4]/p[2]').text.replace('-', '')
        print(f'cep : {cep}')
        sheet.cell(row = start_row, column = 19).value = cep
    except:
        pass
    try:
        check_natureza_juridicas = driver.find_elements(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[1]/div')
        for check_natureza_juridica in check_natureza_juridicas:
            find_natureza_juridica_title = check_natureza_juridica.text.split('\n')
            if find_natureza_juridica_title[0] == "Natureza Jurídica":
                natureza_juridica = find_natureza_juridica_title[1].split(' - ')
                print(f'natureza juridica : {natureza_juridica[1]}')
                sheet.cell(row = start_row, column = 20).value = natureza_juridica[1]
    except:
        pass
    try:
        atividade_principal_cnae = driver.find_element(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[5]/div[1]/p[2]').text.split(' - ')
        print(f'atividade principal cnae : {atividade_principal_cnae[0]}')
        sheet.cell(row = start_row, column = 21).value = atividade_principal_cnae[0]
    except:
        pass
    try:
        check_atividade_secundaria_cnaes = driver.find_element(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[5]/div[2]/p[1]').text
        if check_atividade_secundaria_cnaes == 'Atividades Secundárias':
            atividade_secundaria_cnaes = driver.find_elements(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[5]/div[2]/p')
            cnae_output = []
            text_output = []
            for atividade_secundaria_cnae in atividade_secundaria_cnaes[1:]:
                pre_output = atividade_secundaria_cnae.text.split(' - ')
                cnae_output.append(pre_output[0])
                text_output.append(pre_output[1])
            print(f'atividade_secundaria_cnae : {", ".join(cnae_output)}')
            print(f'atividade_secundaria_text : {", ".join(text_output)}')
            sheet.cell(row = start_row, column = 22).value = ", ".join(cnae_output)
            sheet.cell(row = start_row, column = 23).value = ", ".join(text_output)
    except:
        pass
    try:
        quadro_societários = driver.find_elements(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[4]/div/p')
        quadro_output = []
        for quadro_societário in quadro_societários[1:]:
            quadro_output.append(quadro_societário.text)
        print(f'quadro_societários : {", ".join(quadro_output)}')
        sheet.cell(row = start_row, column = 24).value = ", ".join(quadro_output)
    except:
        pass
    try:
        email = driver.find_element(By.XPATH, '//*[@id="__nuxt"]/div/div[2]/section[1]/div/div/div[4]/div[1]/div[3]/div[2]/p[2]/a').text
        print(email)
        sheet.cell(row = start_row, column = 25).value = email
    except:
        pass
    wb.save('output.xlsx')
    start_row += 1
    driver.delete_all_cookies()
    sleep(1)